
                     # ------------- RESTRICTED MASTER PROBLEM + PRICING PROBLEM ------------- 

from __future__ import print_function

import sys
import time

import cplex
from cplex.exceptions import CplexSolverError
import numpy as np
import openpyxl


# ------------- DATA ------------- (put them in another file)
s_time_loading_data = time.time()

nArcs = 48 # Number of arcs
nStations = 20 # Number of stations
nCommodities = 202 # Number of commodity

#NodeName - NodeID - NodeType - NodeCapacity
node  = [   ['Verona'			,1	,3	,5000],
	    ['Brescia'			,2	,1	,3000],
	    ['Trento'			,3	,1	,3000],
	    ['Bolzano'			,4	,1	,3000],
	    ['Bologna'			,5	,1	,3000],
	    ['Vicenza'			,6	,1	,5000],
	    ['Rovigo'			,7	,1	,5000],
	    ['Padova'			,8	,3	,5000],
	    ['Castelfranco'		,9	,1	,1000],
	    ['Treviso'			,10	,1	,1000],
	    ['Venezia_Mestre'		,11	,1	,5000],
	    ['Venezia_Marghera'		,12	,1	,1000],
	    ['Portogruaro'		,13	,1	,3000],
	    ['Conegliano'		,14	,1	,1000],
	    ['Belluno'			,15	,1	,1000],
	    ['Pordenone'		,16	,2	,3000],
	    ['Udine'			,17	,1	,3000],
	    ['Monfalcone'		,18	,2	,3000],
	    ['Trieste'			,19	,1	,3000],
	    ['Tarvisio'			,20	,1	,3000] 	];

   #FromNode - ToNode - Length - Type (NodeId start at 1)
arc  = [    [1	,2  ,65	,4], #A1
            [1  ,3  ,95 ,3], #A2
            [1  ,5  ,120,2], #A3
            [1  ,6  ,55 ,4], #A4
            [3  ,4  ,55 ,3], #A5
            [6  ,9  ,35 ,2], #A6
            [6  ,8  ,40 ,4], #A7
            [8  ,9  ,30 ,1], #A8
            [10 ,9  ,25 ,2], #A9
            [10 ,11 ,25 ,2], #A10
            [12 ,11 ,5  ,2], #A11
            [8  ,11 ,30 ,4], #A12
            [8  ,7  ,75 ,3], #A13
            [7  ,5  ,45 ,3], #A14
            [13 ,11 ,65 ,3], #A15
            [14 ,10 ,30 ,2], #A16
            [14 ,15 ,50 ,2], #A17
            [14 ,16 ,30 ,2], #A18
            [17 ,16 ,50 ,2], #A19
            [13 ,16 ,30 ,3], #A20
            [13 ,18 ,60 ,3], #A21
            [17 ,18 ,50 ,2], #A22
            [19 ,18 ,35 ,2], #A23
            [17 ,20 ,90 ,3], #A24
            [2  ,1  ,65 ,4], #A25       
            [3  ,1  ,95 ,3], #A26 
            [5  ,1  ,120,2], #A27
            [6  ,1  ,55 ,4], #A28
            [4  ,3  ,55 ,3], #A29
            [9  ,6  ,35 ,2], #A30
            [8  ,6  ,40 ,4], #A31
            [9  ,8  ,30 ,1], #A32
            [9  ,10 ,25 ,2], #A33
            [11 ,10 ,25 ,2], #A34
            [11 ,12 ,5  ,2], #A35
            [11 ,8  ,30 ,4], #A36            
            [7  ,8  ,75 ,3], #A37
            [5  ,7  ,45 ,3], #A38
            [11 ,13 ,65 ,3], #A39 
            [10 ,14 ,30 ,2], #A40
            [15 ,14 ,50 ,2], #A41
            [16 ,14 ,30 ,2], #A42
            [16 ,17 ,50 ,2], #A43
            [16 ,13 ,30 ,3], #A44
            [18 ,13 ,60 ,3], #A45
            [18 ,17 ,50 ,2], #A46
            [18 ,19 ,35 ,2], #A47
            [20 ,17 ,90 ,3] ]; #A48

# From - To - Quantity (NodeId starts at 0) 
commodities = [
	[14, 18, 5],  [14, 19, 6],  [14, 1, 11],  [14, 4, 18],  
	[4, 5, 65],  [4, 0, 145],  [4, 11, 40],  [4, 11, 54],  [4, 11, 12],  
	[4, 16, 31],  [4, 18, 48],  [4, 18, 5],  [4, 9, 40],  [4, 2, 28],  
	[4, 19, 50],  [4, 19, 52],  [4, 19, 35],  [4, 6, 21],  [4, 12, 30],  
	[4, 15, 36],  [4, 7, 96],  [4, 17, 25],  [4, 8, 31],  [4, 1, 43],  
	[4, 3, 10],  [4, 3, 12],  [4, 3, 25],  [4, 3, 66],  [4, 3, 80],  
	[4, 14, 13],  [3, 5, 40],  [3, 0, 128],  [3, 11, 106],  [3, 16, 25],  
	[3, 18, 18],  [3, 9, 20],  [3, 2, 26],  [3, 6, 6],  [3, 12, 20],  
	[3, 15, 18],  [3, 7, 118],  [3, 17, 10],  [3, 13, 10],  [3, 8, 10],  
	[3, 1, 85],  [3, 1, 25],  [3, 1, 55],  [3, 1, 25],  [3, 4, 20],  
	[3, 4, 24],  [3, 4, 101],  [1, 5, 36],  [1, 0, 80],  [1, 0, 62],  
	[1, 11, 111],  [1, 16, 41],  [1, 18, 15],  [1, 18, 55],  [1, 9, 25],  
	[1, 2, 20],  [1, 19, 160],  [1, 12, 25],  [1, 15, 16],  [1, 7, 24],  
	[1, 7, 44],  [1, 7, 59],  [1, 17, 21],  [1, 13, 26],  [1, 8, 21],  
	[1, 3, 25],  [1, 3, 44],  [1, 3, 6],  [1, 3, 68],  [1, 3, 20],  
	[1, 4, 36],  [1, 14, 50],  [8, 19, 38],  [8, 1, 21],  [8, 3, 13],  
	[8, 4, 46],  [13, 19, 70],  [13, 1, 28],  [13, 3, 10],  [13, 4, 15],  
	[17, 5, 10],  [17, 0, 15],  [17, 18, 16],  [17, 19, 25],  [17, 1, 55],  
	[17, 3, 36],  [17, 4, 33],  [7, 18, 45],  [7, 2, 6],  [7, 19, 60],  
	[7, 1, 130],  [7, 1, 22],  [7, 1, 15],  [7, 3, 26],  [7, 3, 24],  
	[7, 3, 40],  [7, 3, 15],  [7, 4, 105],  [15, 18, 31],  [15, 19, 51],  
	[15, 1, 36],  [15, 3, 41],  [15, 4, 30],  [15, 4, 25],  [12, 0, 6],  
	[12, 18, 15],  [12, 19, 33],  [12, 1, 33],  [12, 3, 16],  [12, 4, 30],  
	[6, 18, 11],  [6, 19, 23],  [6, 1, 43],  [6, 3, 21],  [6, 4, 58],  
	[19, 5, 51],  [19, 0, 12],  [19, 0, 60],  [19, 11, 78],  [19, 16, 53],  
	[19, 9, 31],  [19, 6, 10],  [19, 12, 15],  [19, 15, 46],  [19, 7, 73],  
	[19, 17, 30],  [19, 13, 26],  [19, 8, 21],  [19, 1, 40],  [19, 1, 50],  
	[19, 1, 18],  [19, 4, 60],  [19, 4, 50],  [19, 14, 8],  [2, 11, 5],  
	[2, 18, 10],  [2, 1, 25],  [2, 3, 21],  [2, 4, 15],  [9, 0, 11],  
	[9, 18, 16],  [9, 19, 30],  [9, 1, 55],  [9, 3, 26],  [9, 4, 43],  
	[18, 5, 50],  [18, 0, 35],  [18, 11, 21],  [18, 16, 26],  [18, 9, 10],  
	[18, 2, 6],  [18, 6, 5],  [18, 12, 6],  [18, 15, 10],  [18, 7, 46],  
	[18, 17, 8],  [18, 1, 13],  [18, 1, 45],  [18, 3, 20],  [18, 4, 43],  
	[18, 14, 8],  [16, 5, 5],  [16, 0, 13],  [16, 18, 28],  [16, 19, 56],  
	[16, 1, 63],  [16, 3, 18],  [16, 4, 45],  [16, 4, 33],  [11, 0, 8],  
	[11, 18, 43],  [11, 2, 11],  [11, 19, 51],  [11, 1, 20],  [11, 1, 55],  
	[11, 1, 15],  [11, 1, 52],  [11, 3, 81],  [11, 4, 30],  [11, 4, 70],  
	[0, 11, 20],  [0, 16, 13],  [0, 18, 40],  [0, 19, 83],  [0, 17, 30],  
	[0, 1, 20],  [0, 1, 65],  [0, 1, 70],  [0, 3, 45],  [0, 3, 80],  
	[0, 4, 123],  [5, 16, 10],  [5, 18, 43],  [5, 19, 46],  [5, 17, 13],  
	[5, 1, 98],  [5, 3, 70],  [5, 4, 73] 
]

scale_capacity = 2.4 # Variable used to scale up or down the capacity of the arcs (depending on their type)
capacity_arc = [200,350,500,650] # Define a capacity value for each type of arcs
capacity_arc = [cap*scale_capacity for cap in capacity_arc] # Scale them by a constant

capacity_node = [1000,2000,3000] # Define a capacity value for each type of nodes
capacity_node = [cap*scale_capacity for cap in capacity_node] # Scale them by a constant

cost = np.zeros((nStations, nStations, 2), dtype=int) # Represent the cost between each node with the corresponding arc_id, if equals 0 then no arc between the two nodes, 

#Read the data file
excelfile = openpyxl.load_workbook('/Users/nicolas/Documents/EPFL/4.Printemps2018/Project/Small_Instance/SemesterProject_MFP_SmallInstance.xlsx')  # open a excel file with .xlsx format
sheet2 = excelfile.get_sheet_by_name("Arcs_Final")

# Fill the cost variable by reading the arc cost and index in the file
for row_index in range(2, (sheet2.max_row+1)):
    cost[sheet2.cell(row=row_index, column=3).value - 1, sheet2.cell(row=row_index, column=4).value - 1] = [sheet2.cell(row=row_index, column=5).value, row_index-2]

def dijkstra(origin):

    # -------- DIJKSTRA'S ALGORITHM --------
    
    labels = np.full((nStations), np.inf) # Contains the cost of the shortest paths from origin with init value = infinity
    labels[origin] = 0
    toTreat = np.array([origin]) # Node that need to be processed
    
    while True:
        currentNode = toTreat[np.argmin(np.take(labels, toTreat))] # Pick the node with lowest shortest path from the set toTreat
        for i in range(nStations):
            if(cost[currentNode, i, 0] > 0): #Among all neighbors of currentNodes
                if(labels[i] > labels[currentNode] + cost[currentNode, i, 0]): # If dual constraint violated
                    
                    labels[i] = labels[currentNode] + cost[currentNode, i, 0] # Lower it to satisfy the feasibility and slackness property
                    toTreat = np.append(toTreat, i) # As we found a better path to go to i, need to update the neighbors of j

        toTreat = np.delete(toTreat,np.where(toTreat==currentNode)[0]) # Delete the node we just processed
        if(toTreat.size == 0): # If there is no node left to treat, we are sure we obtained the best solution.
            break
    return labels

all_Arc = [] # Will contain the arcs of each path found in the initial set
all_Cost = [] # Will contain the cost of each path in all_Arc

def getAllPath(i,j, optimal): #Get all paths from i to j with length < limit * optimal
    global all_Arc, all_Cost
    
    # -------- 1/2 GET ALL PATHS FROM i TO j --------
    
    all_Arc = []
    all_Cost = []
    visited =[False]*nStations #Keep track of the visited node to avoid cycle
    arc = [0] * nArcs # Write the current path in it (which arcs it contains)
    printAllPath(i, -1 ,j, visited, arc, 0, optimal)
    
    for i in range(len(all_Arc)):
        all_Arc[i].append(all_Cost[i])

# Recursive method based on Depth First Search 

# Algorithm based on Depth-First-Search : Find all paths from i to j,
# (precedent is the last node we processed
# ,arc contains the path in process
#, costP = current cost of the path in process
#, optimal = shortest path distance between i and j)
#
#Basically, explore all possible paths from i, if we reach j, append the path to solution.
#
# Start at node i, mark node i as visited,
# If i == j : We succesfully reached destination, append the pathway to the solution and its cost to costP
#
# Else :    We recursivey iterate over all the unvisited neighbours k of i, which are not too far, using the adjency matrix (Depth First Search)
#           and update the arc(i, k) to 1

# Finally : Each time we are done with a path we reset arc, costP and visited variables

def printAllPath(i, precedent , j, visited, arc, costP, optimal):
    global  all_Cost, all_Arc
    
    # -------- 2/2 THROUGH A RECURSIVE SEARCH FUNCTION --------
    
    visited[i] = True #To remember that we visited current node
    if(precedent > -1): 
        arc[cost[precedent, i, 1]] = 1.0 # If you moved from a previous node, add the arc between them to the potential solution path
    
    if(i==j): #If we reach the destination, we print the path
        all_Cost.append(costP) # Add the cost of the succesfull path
        all_Arc.append(arc[:]) # Add the path to global solution
        
    else: #Otherwise we iterate over all neighbors unvisited nodes (by depth), only if the future distance is not too far from the shortest path 
        for k in range(nStations):
            if(cost[i, k, 0] > 0 and visited[k]==False and (costP+cost[i, k, 0]) <= limit*optimal) :
                printAllPath(k, i ,j ,visited, arc ,costP + cost[i, k, 0], optimal)
            
    if(precedent > -1): # Once you reached a leaf (no where to go), reset the cost, arc and visited variable to their previous value
        costP = costP - cost[precedent, i, 0]
        arc[cost[precedent, i, 1]] = 0
    visited[i] = False

def getInitSolution_Dijkstra():

    # ------------- GET INITIAL SET OF PROMISSING VARIABLES -------------

    shortest_paths = np.zeros((nStations, nStations)) # Represent the shortest path distance between each nodes

    #Run Dijkstra's Algo for each node to get the shortest path distance
    for i in range(nStations-1): #Don't need the last node, if you have already find all the shortest path from every node except the last, you already computed the info for the last node
        shortest_paths[i] = dijkstra(i)
        shortest_paths[:, i] = shortest_paths[i] # Fill the matrix row-wise and column-wise, as the length is symetric

    pathsK = [[[]] for i in range(nCommodities)] # Contains all initial paths for each commodity
    
    #pathsK[k,i,j] with k = commodity_id // i = path_id of commodity k // j = arc_id of the path 
    
    #Now obtain for each commodity, obtain all possible path with length < CST * OPT_Distance
    for i in range(nCommodities):
        getAllPath(commodities[i][0], commodities[i][1], shortest_paths[commodities[i][0], commodities[i][1]] )
        pathsK[i] = all_Arc # Put all the result in pathsK variable
        
    return pathsK

def dual_of_Restrited(pathsK): # To obtain dual variables

    # ------------- DUAL OF THE RESTRICTED MASTER PROBLEM -------------
    
    model = cplex.Cplex() # Initialize the model
    model.objective.set_sense(model.objective.sense.maximize) ## Say that we want to maximize the objective function

    #Add variables
    for i in range(nCommodities):
        model.variables.add(obj = [1], lb = [-cplex.infinity], ub = [cplex.infinity],#obj contains the coefficients of the decision variables in the objective function
                            types = [model.variables.type.continuous],
                            names = ['Y( K_' + str(i) + ')'])

    for i in range(nArcs):
        model.variables.add(obj = [ capacity_arc[arc[i][3]-1] ], lb = [-cplex.infinity], ub = [0],#obj contains the coefficients of the decision variables in the objective function
                            types = [model.variables.type.continuous],
                            names = ['Y( A_' + str(arc[i][0]-1) + '_' + str(arc[i][1]-1) + ')'])
        
    for i in range(nStations):
        model.variables.add(obj = [ capacity_node[node[i][2]-1] ], lb = [-cplex.infinity], ub = [0],#obj contains the coefficients of the decision variables in the objective function
                            types = [model.variables.type.continuous],
                            names = ['Y( ' + str(node[i][0])  + ')'])
        

    #Add constraints
    for i in range(nCommodities):
        for j in range(len(pathsK[i])):
            ind = []# Put the indices of the non-zero variables of the current constraint
            val = []# Put their coefficients here
            row = []
            ind.append(i)
            val.append(1)
            for k in range(nArcs):
                if(pathsK[i][j][k] == 1):
                    ind.append(nCommodities+k)
                    val.append(commodities[i][2])
            for k in range(nStations):
                for l in range(nArcs):
                    if(pathsK[i][j][l] == 1 and arc[l][1]-1 == k):
                        ind.append(nCommodities+nArcs+k)
                        val.append(commodities[i][2])
            row.append([ind, val])
            model.linear_constraints.add(lin_expr = row,
                                     senses = "L", #Equality constraint
                                     rhs = [ pathsK[i][j][nArcs]*float(commodities[i][2]) ] ) # Right Hand Side of the constraint



    try:
        print("\n\n-----------RESTRICTED DUAL SOLUTION : -----------\n")
        model.solve()
        model.write('test3.lp')
    except CplexSolverError as e:
        print("Exception raised during restricted master problem: " + e)

    return model.solution.get_values()[:nCommodities], model.solution.get_values()[nCommodities:nCommodities + nArcs], model.solution.get_values()[nArcs + nCommodities:]
    # Return the dual variables corresponding to the commodities and arcs constraints

def restricted_Master(pathsK):
    
    # ------------- SOLVE THE RESTRICTED MASTER PROBLEM -------------
        
    model = cplex.Cplex() # Initialize the model
    model.objective.set_sense(model.objective.sense.minimize) ## Say that we want to minimize the objective function
    
    #Create decision variables
    for i in range(nCommodities):
        for j in range(len(pathsK[i])):
            model.variables.add(obj = [pathsK[i][j][nArcs]*float(commodities[i][2])], lb = [0], ub = [1],#obj contains the coefficients of the decision variables in the objective function                               
                                types = [model.variables.type.continuous],
                                names = ['P(' + str(i) + ',' + str(j) + ')']) 

    #Add constraints
            
    #Flow conservation constraints :
    count = 0 # To iterate over the index of the decision variables
    for i in range(nCommodities):
        ind = [] # Put the indices of the non-zero variables of the current constraint
        val = [] # Put their coefficients here
        row = []
        for j in range(len(pathsK[i])):
            ind.append(count) # Give the indices of the variables contained in the contraint
            val.append(1) # With their coefficient
            count += 1
        row.append([ind, val])
        model.linear_constraints.add(lin_expr = row,
                                     senses = "E", #Equality constraint
                                     rhs = [1] ) # Right Hand Side of the constraint
        
    #Arc capacity constraints :
    for i in range(nArcs): #For each arc
        ind, val, row = [], [], []
        count = 0
        for j in range(nCommodities): #For each commodity paths, check each time a path contains the arc, 
            for k in range(len(pathsK[j])):
                if(pathsK[j][k][i] == 1): # If it is the case, add the decision variable index to the constraint
                    ind.append(count)
                    val.append(commodities[j][2]) # With its coefficiant
                count += 1
        row.append([ind, val])
        model.linear_constraints.add(lin_expr = row,
                                     senses = "L", # Less-than
                                     rhs = [ capacity_arc[arc[i][3]-1] ] ) #Capacity of the arc (-1 because type_id start at 1)

    #Node capacity constraints :
    for i in range(nStations): #For each node
        ind, val, row = [], [], []
        count = 0
        for j in range(nCommodities): #For each commodity paths, check each time a path contains the node (do not include starting node)
            for k in range(len(pathsK[j])):
                for l in range(nArcs):
                    if(pathsK[j][k][l] == 1 and arc[l][1]-1 == i): # If it is the case, add the decision variable index to the constraint
                        ind.append(count)
                        val.append(commodities[j][2]) # With its coefficiant
                count += 1
        row.append([ind, val])
        model.linear_constraints.add(lin_expr = row,
                                     senses = "L", # Less-than
                                     rhs = [ capacity_node[node[i][2]-1] ] ) #Capacity of the node (-1 because type_id start at 1)

    

    try:
        print("\n\n-----------RESTRICTED MASTER SOLUTION : -----------\n")
        model.solve()
        model.write('test.lp')
        print("\n")
        print("Solution primal : ",model.solution.get_values())
        #If the try statement did well
        count = 0
        for i in range(nCommodities):
            for j in range(len(pathsK[i])):
                indices = [k for k, x in enumerate(pathsK[i][j]) if x == 1] #Get the indices of the arcs contained in the current path
                print("\t", model.solution.get_values(count)*commodities[i][2] ,"quantity of commodity n°", i ,"on path", node[commodities[i][0]][0],''
                      + ' '.join([node[arc[k][1]-1][0] for k in indices]) + ". Length path : " + str(pathsK[i][j][nArcs]) )
                count += 1
        print("\nTotal cost = " + str(model.solution.get_objective_value()))

        dualCommodities = [] #Contain the dual variables corresponding to the flow conservation constraints in the primal
        dualArcs = [] #Contain the dual variables corresponding to the arc capacity constraints in the primal
        dualStations = [] #Contain the dual variables corresponding to the node capacity constraints in the primal
        
        dualCommodities, dualArcs, dualStations = dual_of_Restrited(pathsK) # Compute them by solving the dual
        #Do it in another try except bloc
 #       dual = model.solution.get_dual_values()
        
#        dualCommodities = dual[:nCommodities]
#        print("commot", dualCommodities)
#        dualArcs = dual[nCommodities:nArcs+nCommodities]
#        print("arc", dualArcs)
#        dualStations = dual[nArcs+nCommodities:nArcs+nCommodities+nStations]
#        print("station", dualStations)

  
        for i in range(len(dualCommodities)):
            print("\nDual values y_K" + str(i+1) + " = "+ str(dualCommodities[i]))
        for i in range(len(dualArcs)):
            print("Dual values y_Arc" + str(i+1) + " = "+ str(dualArcs[i]))
        for i in range(len(dualStations)):
            print("Dual values y_Node" + str(i+1) + " = "+ str(dualStations[i]))
            
    except CplexSolverError as e:
        print("Exception raised during restricted master problem: ", e)
        return [-1] * 5 # return -1 to indicate the infeasibility of restricted master problem

    return model.solution.get_objective_value(), model.solution.get_values(), dualCommodities, dualArcs, dualStations

def pricingProblem(dualCommodities, dualArcs, dualStations): # For the moment, return the path corresponding to the most violated constraint

    # ------------- SOLVE THE PRICING PROBLEM-------------
    
    reducedCost = 0
    forCommodity = 0
    bestPath = []
    for i in range(nCommodities): #Solve the shortest path problem (with updated length) for each commodity
        model = cplex.Cplex()
        model.objective.set_sense(model.objective.sense.minimize) ## Say that we want to minimize the objective function
    
        #Create decision variables (array of nArcs size)
        for j in range(nArcs):
            model.variables.add(obj = [commodities[i][2]*(arc[j][2] - dualArcs[j] - dualStations[ arc[j][1]-1 ] )], #obj contains the coefficients of the decision variables in the objective function
                                lb = [0], ub = [1],
                                types = [model.variables.type.integer],
                                names = ['alpha ( ' + str(j) + ' )'])
            
            
        model.objective.set_offset(-dualCommodities[i]) # Check if does what we want

        #Add the consistency constraint ( alpha in P(k) )
        #For each node, check if it is a starting node, ending node or in-between node
        for k in range(nStations):
            ind = [] 
            val = [] 
            row = []
            if(commodities[i][0] != k and commodities[i][1] != k ): #If the node is in between 
                rhs = [0]
                for j in range(nArcs):
                    if(arc[j][0]-1 == k):# Compute its leaving flow (-1 because in the data we start with index 1 and not 0)
                        ind.append(j) 
                        val.append(1) 
                    elif(arc[j][1]-1 == k): # Minus what comes in
                        ind.append(j)
                        val.append(-1)
            elif(commodities[i][0] == k): # If the node is the starting node of the commodity
                rhs = [1]
                for j in range(nArcs):
                    if(arc[j][0]-1 == k):# Compute its leaving flow
                        ind.append(j) 
                        val.append(1) 
                    elif(arc[j][1]-1 == k): # Minus what comes in
                        ind.append(j)
                        val.append(-1)
            elif(commodities[i][1] == k): # If the node is the starting node of the commodity
                rhs = [-1]
                for j in range(nArcs):
                    if(arc[j][0]-1 == k):# Compute its leaving flow
                        ind.append(j) 
                        val.append(1) 
                    elif(arc[j][1]-1 == k): # Minus what comes in
                        ind.append(j)
                        val.append(-1)                       
            row.append([ind, val])
            model.linear_constraints.add(lin_expr = row,
                                             senses = "E", #Equality constraint
                                             rhs = rhs )
            
        try:
            print("\n\n-----------PRICING PROBLEM SOLUTION FOR COMMODITY n°", i ,": -----------\n")
            model.solve()
            model.write('test2.lp')
            print()
            print("\n\tREDUCED COST : ", model.solution.get_objective_value())
            print("\n\tNEW PATH : ", model.solution.get_values())
        except CplexSolverError as e:
            print("Exception raised during pricing problem: " + e)
        else:
            if(model.solution.get_objective_value() < reducedCost): # If we obtained a more violated constraint, take it
                reducedCost = model.solution.get_objective_value()
                bestPath = model.solution.get_values()
                forCommodity = i
        
    if(reducedCost < 0 ): #Compute the cost of the new path
        tempCost = 0
        for i in range(nArcs):
            if(bestPath[i] == 1):
                tempCost += arc[i][2]
        bestPath.append(tempCost) #Put the cost of the path at the bottom of its arcs description
       
    return reducedCost, bestPath, forCommodity
        
            

# --- BEGIN HERE ---

limit = 1.6 # Include all paths 'p' satisfying dist(p) < limit*shortest_path

def getInitSet():
    pathsK = getInitSolution_Dijkstra()
    return pathsK

t_time_loading_data = time.time()

pathsK = getInitSet() # Get the initial set of variable through Dijkstra and "< limit*shortest_path " method

while True:
    #Check if init set is feasible
    while True:
        print('VALUE LIMIT : ', limit)
        obj_Function, solution, dualCommodities, dualArcs, dualStations = restricted_Master(pathsK) #Solve the restricted master problem
        if(obj_Function > -1):
            break
        limit *= 1.05 # Increase by 5%
        pathsK = getInitSet()
        
    t_time_init_set = time.time()
    #Then iterate over pricing and restricted master problem
    reducedCost, newPath, forCommodity = pricingProblem(dualCommodities, dualArcs, dualStations) #Solve the pricing problem
    
    if(round(reducedCost) == 0): # round to avoid minor computational error (10^-23 for 0)
        t_time_solving = time.time()
        count = 0
        #Print final solution
        for i in range(nCommodities):
            print()
            for j in range(len(pathsK[i])):
                
                indices = [k for k, x in enumerate(pathsK[i][j]) if x == 1] #Get the indices of the arcs contained in the current path
                print("\t", solution[count]*commodities[i][2] ,"quantity of commodity n°", i+1 ,"on path", node[commodities[i][0]][0],''
                      + ' '.join([node[arc[k][1]-1][0] for k in indices]) + ". Length path : " + str(pathsK[i][j][nArcs]) )
                count += 1
        print("\nTotal cost = " + str(obj_Function))
        print("\nLoading data duration : ", t_time_loading_data - s_time_loading_data)
        print("Finding initial set duration : ", t_time_init_set - t_time_loading_data)
        print("Solving problem duration : ", t_time_solving - t_time_init_set)
        print("limit : ", limit)
        
        
        break
    else:
        pathsK[forCommodity].append(newPath) # We iterate with a new path for commodity n° "forCommodity"



